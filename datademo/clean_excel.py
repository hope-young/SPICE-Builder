"""
clean_excel.py
==============
将 Silicon-Magic 半导体测试报告 Excel 文件清洗为标准 CSV。

输入: datademo/testdatafrom1506.xlsx
输出: datademo/cleaned/*.csv (每条曲线一个 CSV)
      datademo/inspect/cleaning_report.json (清洗报告)

数据 sheet 列表:
  - ID-VGS_5VDS, ID-VGS_0.5VDS  → Id-Vg (多温度)
  - ID-VDS                       → Id-Vd
  - VGS-Qg                       → 栅电荷
  - CissCossCrss-VDS             → C-V
  - CissCrss-VGS                 → C-V vs Vgs
  - IS-VSD                       → 体二极管
  - TJ_DC                        → 温度 DC 参数 (BVDSS/Vth/RDS/VSD)
  - RDSon-VGS                    → Rdson vs Vgs
  - ID-VTH                       → 阈值特性
  - Qrr                          → 体二极管反向恢复
"""
import json
import warnings
from pathlib import Path
from openpyxl import load_workbook

warnings.filterwarnings('ignore')

INPUT_FILE = Path(__file__).parent / "testdatafrom1506.xlsx"
OUTPUT_DIR = Path(__file__).parent / "cleaned"
INSPECT_DIR = Path(__file__).parent / "inspect"
OUTPUT_DIR.mkdir(exist_ok=True)
INSPECT_DIR.mkdir(exist_ok=True)


def _is_numeric(v):
    """检查是否可转为 float（排除 None/NaN/#N/A 等）"""
    if v is None or v == '':
        return False
    if isinstance(v, str):
        s = v.strip()
        if s in ('', 'NaN', 'N/A', '#N/A', '#REF!', '#DIV/0!', '#VALUE!', '-Infinity', 'Infinity'):
            return False
        try:
            float(s)
            return True
        except (ValueError, TypeError):
            return False
    return isinstance(v, (int, float))


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s in ('', 'NaN', 'N/A', '#N/A', '#REF!', '#DIV/0!', '#VALUE!'):
            return None
        if s in ('-Infinity',):
            return None
        if s == 'Infinity':
            return None
        try:
            return float(s)
        except (ValueError, TypeError):
            return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _find_header_row(ws, max_scan=10, min_numeric_cols=3):
    """扫描前 N 行，找到第一个有 ≥min_numeric_cols 个数值列的行（视为数据起始行）"""
    for r in range(1, min(max_scan + 1, ws.max_row + 1)):
        row = [c.value for c in ws[r]]
        n_numeric = sum(1 for v in row if _is_numeric(v))
        if n_numeric >= min_numeric_cols:
            return r
    return None


def _strip_units(val, hint_unit=None):
    """简单单位转换：mA→A, nF→F, nC→C, uA→A 等"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if not isinstance(val, str):
        return None
    s = val.strip()
    f = _to_float(s)
    if f is None:
        return None
    return f


# ============================================================
#  Sheet parsers (每个 sheet 一个)
# ============================================================

def parse_idvg_5vds(ws, wb):
    """ID-VGS_5VDS: Id-Vg @ Vds=5V, 多温度
    列布局: VGS | ID(-55) | ID(25) | ID(150) | gfs(-55) | gfs(25) | gfs(150)
    """
    # 数据从第 2 行开始（第一行是 metadata）
    # 实际列: G=VGS, H,I,J=-55度id, K,L,M=25度id, ...
    # 看 row 1: 1, 'VGS', 25, None, 150 → 列 G=1, H='VGS', I=25, K=150
    # Row 2: 1, 0, 0.0287..., ... → G=1, H=0(VGS), I=Id@25, K=Id@150

    # 实际位置: G=0, H=0 → 但前面有 'ID-VGS' 标识
    # 简化: 找 VGS 列 + 3 个 ID 列
    rows = []
    for r in range(2, ws.max_row + 1):
        row_vals = [c.value for c in ws[r]]
        # VGS 在 col 7 (G), ID-55 在 8, ID-25 在 9, ID-150 在 10
        vgs = _to_float(row_vals[6]) if len(row_vals) > 6 else None
        id_neg55 = _to_float(row_vals[7]) if len(row_vals) > 7 else None
        id_25 = _to_float(row_vals[8]) if len(row_vals) > 8 else None
        id_150 = _to_float(row_vals[9]) if len(row_vals) > 9 else None
        if vgs is None:
            continue
        # 至少有一个有效 Id
        if id_neg55 is None and id_25 is None and id_150 is None:
            continue
        rows.append({
            'vgs_v': vgs,
            'id_neg55_a': id_neg55,
            'id_25c_a': id_25,
            'id_150c_a': id_150,
            'vds_v': 5.0,
        })
    return rows


def parse_idvg_05vds(ws, wb):
    """ID-VGS_0.5VDS: Id-Vg @ Vds=0.5V, 多温度
    同样的列结构
    """
    rows = []
    for r in range(2, ws.max_row + 1):
        row_vals = [c.value for c in ws[r]]
        vgs = _to_float(row_vals[6]) if len(row_vals) > 6 else None
        id_neg55 = _to_float(row_vals[7]) if len(row_vals) > 7 else None
        id_25 = _to_float(row_vals[8]) if len(row_vals) > 8 else None
        id_150 = _to_float(row_vals[9]) if len(row_vals) > 9 else None
        if vgs is None:
            continue
        if id_neg55 is None and id_25 is None and id_150 is None:
            continue
        rows.append({
            'vgs_v': vgs,
            'id_neg55_a': id_neg55,
            'id_25c_a': id_25,
            'id_150c_a': id_150,
            'vds_v': 0.5,
        })
    return rows


def parse_idvd(ws, wb):
    """ID-VDS: Id-Vd, 多 Vgs
    列布局: VDS | ID(Vgs=2.5) | ID(Vgs=3) | ID(Vgs=3.5) | ...
    """
    # 看 row 1: VGS = 2.5V, VGS = 3V, VGS = 3.5V ...
    # Row 2: VDS, ID, RDS(Ω), RDS(mΩ), Normalized, VDS, ID, RDS(Ω), ...
    # Row 3 开始数据
    # 找 VGS 标签
    row1 = [c.value for c in ws[1]]
    row2 = [c.value for c in ws[2]]
    vgs_cols = []  # [(col_idx, vgs_value), ...]
    for c in range(len(row1)):
        v = row1[c]
        if isinstance(v, str) and 'VGS' in str(v):
            # 提取 Vgs 数字
            import re
            m = re.search(r'([\d.]+)\s*V', str(v))
            if m:
                vgs_cols.append((c, float(m.group(1))))
    # 也可能只有 row2 是 'VDS','ID' 列标识
    if not vgs_cols:
        # fallback: 从 row1 找 "VGS = X V" 模式
        for c in range(len(row1)):
            v = row1[c]
            if v is not None:
                import re
                m = re.search(r'VGS\s*=\s*([\d.]+)\s*V', str(v))
                if m:
                    vgs_cols.append((c, float(m.group(1))))

    # 每个 Vgs group: VDS col, ID col, RDS col
    rows = []
    for r in range(3, ws.max_row + 1):
        row_vals = [c.value for c in ws[r]]
        for col_idx, vgs_v in vgs_cols:
            vds = _to_float(row_vals[col_idx]) if col_idx < len(row_vals) else None
            id_v = _to_float(row_vals[col_idx + 1]) if (col_idx + 1) < len(row_vals) else None
            if vds is None or id_v is None:
                continue
            if vds <= 0:
                continue
            rows.append({
                'vds_v': vds,
                'id_a': id_v,
                'vgs_v': vgs_v,
                'temperature_c': 25,  # 默认
            })
    return rows


def parse_vgs_qg(ws, wb):
    """VGS-Qg: 栅电荷曲线
    列布局: Vgs | Qg(nC) | Ciss(eff) | Id
    """
    # 看 row 1: Qg, Vgs, Qg (nC), Ciss(eff), Id
    # Row 2 开始数据
    # 找 Vgs, Qg 列
    row1 = [c.value for c in ws[1]]
    vgs_col = None
    qg_col = None
    cis_col = None
    id_col = None
    for c, v in enumerate(row1):
        if v is None:
            continue
        sv = str(v).lower()
        if vgs_col is None and 'vgs' in sv and 'off' not in sv and 'on' not in sv and 'th' not in sv:
            vgs_col = c
        if qg_col is None and 'qg' in sv and 'n' in sv:
            qg_col = c
        if cis_col is None and 'ciss' in sv:
            cis_col = c
        if id_col is None and sv.strip() == 'id':
            id_col = c
    # 看实际数据：col 14 (O)='Qg', col 15 (P)='Vgs', col 16 (Q)='Qg (nC)', col 17 (R)='Ciss(eff)', col 18 (S)='Id'
    # 强制使用已知列
    vgs_col = 15  # P
    qg_col = 16   # Q (nC)
    cis_col = 17  # R
    id_col = 18   # S

    rows = []
    for r in range(2, ws.max_row + 1):
        row_vals = [c.value for c in ws[r]]
        if vgs_col is None or vgs_col >= len(row_vals):
            continue
        vgs = _to_float(row_vals[vgs_col])
        qg = _to_float(row_vals[qg_col]) if qg_col < len(row_vals) else None
        ciss = _to_float(row_vals[cis_col]) if cis_col < len(row_vals) else None
        id_v = _to_float(row_vals[id_col]) if id_col < len(row_vals) else None
        if vgs is None or qg is None:
            continue
        # 单位: Qg 是 nC
        rows.append({
            'vgs_v': vgs,
            'qg_nc': qg,
            'ciss_pf': ciss * 1e3 if ciss is not None else None,  # F → pF
            'id_a': id_v,
        })
    return rows


def parse_cv_vds(ws, wb):
    """CissCossCrss-VDS: C-V 曲线
    列布局: Vds | Ciss(F) | Ciss(pF) | Vds | Coss(F) | Coss(pF) | Vds | Crss(F) | Crss(pF) | ...
    """
    # 看 row 1: VDS, Ciss(F), Ciss(pF), VDS, Coss(F), Coss(pF), VDS, Crss(F), Crss(pF), Cgd/Ciss, ...
    # Row 2 起数据
    # 列位置: 5=Vds_Ciss, 6=Ciss_F, 7=Ciss_pF, 8=Vds_Coss, 9=Coss_F, 10=Coss_pF, 11=Vds_Crss, 12=Crss_F, 13=Crss_pF
    rows = []
    for r in range(2, ws.max_row + 1):
        row_vals = [c.value for c in ws[r]]
        vds_c = _to_float(row_vals[5]) if len(row_vals) > 5 else None
        ciss_f = _to_float(row_vals[6]) if len(row_vals) > 6 else None
        vds_o = _to_float(row_vals[8]) if len(row_vals) > 8 else None
        oss_f = _to_float(row_vals[9]) if len(row_vals) > 9 else None
        vds_r = _to_float(row_vals[11]) if len(row_vals) > 11 else None
        rss_f = _to_float(row_vals[12]) if len(row_vals) > 12 else None
        # 任何一个有效
        if ciss_f is None and oss_f is None and rss_f is None:
            continue
        rows.append({
            'vds_v': vds_c if vds_c is not None else (vds_o if vds_o is not None else vds_r),
            'ciss_f': ciss_f,
            'coss_f': oss_f,
            'crss_f': rss_f,
        })
    return rows


def parse_isvsd(ws, wb):
    """IS-VSD: 体二极管
    列布局: VSD | IS(-55) | VSD | IS(25) | VSD | IS(150) | VSD | IS(Vgs=0) | ...
    """
    # 看 row 1: -55, None, 25, None, 150, None (3 温度)
    # Row 2: VSD, IS, VSD, IS, VSD, IS (列标识)
    # Row 3 起数据
    # 列: 6=VSD(-55), 7=IS(-55), 8=VSD(25), 9=IS(25), 10=VSD(150), 11=IS(150)
    # 还有 RAWDATA 块: Vgs=0V, Vgs=-1V, Vgs=-2V
    rows = []
    for r in range(2, ws.max_row + 1):
        row_vals = [c.value for c in ws[r]]
        if len(row_vals) < 12:
            continue
        vsd_55 = _to_float(row_vals[6])
        is_55 = _to_float(row_vals[7])
        vsd_25 = _to_float(row_vals[8])
        is_25 = _to_float(row_vals[9])
        vsd_150 = _to_float(row_vals[10])
        is_150 = _to_float(row_vals[11])
        if vsd_55 is not None and is_55 is not None:
            rows.append({
                'vsd_v': abs(vsd_55), 'is_a': abs(is_55), 'temperature_c': -55, 'vgs_v': 0
            })
        if vsd_25 is not None and is_25 is not None:
            rows.append({
                'vsd_v': abs(vsd_25), 'is_a': abs(is_25), 'temperature_c': 25, 'vgs_v': 0
            })
        if vsd_150 is not None and is_150 is not None:
            rows.append({
                'vsd_v': abs(vsd_150), 'is_a': abs(is_150), 'temperature_c': 150, 'vgs_v': 0
            })
    return rows


def parse_tj_dc(ws, wb):
    """TJ_DC: 温度 DC 参数
    列: BVDSS, BVDSS_1, BVDSS_2, BVDSS_3, VGS(th), VGS(th)_1, IGSS, IGSS(-), IDSS, IDSS_1,
        RDS(on), RDS(on)_1, RDS(on)_2, RDS(on)_3, VSD, ...
    行为: -55°C, 25°C, 150°C
    """
    # 列位置: 6=BVDSS, 10=VGS(th), 12=IGSS, 14=IDSS, 16=RDS(on), 20=VSD
    # 看 row 1: BVDSS,BVDSS_1,...,VSD
    # Row 2: 温度=-55, 各参数值
    rows = []
    for r in range(2, ws.max_row + 1):
        row_vals = [c.value for c in ws[r]]
        if len(row_vals) < 6:
            continue
        temp = _to_float(row_vals[4])  # 温度在 col 5 (E)
        if temp is None:
            continue
        rows.append({
            'temperature_c': temp,
            'bvdss_v': _to_float(row_vals[5]) if len(row_vals) > 5 else None,
            'vth_v': _to_float(row_vals[9]) if len(row_vals) > 9 else None,
            'igss_a': _to_float(row_vals[11]) if len(row_vals) > 11 else None,
            'idss_a': _to_float(row_vals[13]) if len(row_vals) > 13 else None,
            'rdson_ohm': _to_float(row_vals[17]) if len(row_vals) > 17 else None,
            'vsd_v': _to_float(row_vals[19]) if len(row_vals) > 19 else None,
        })
    return rows


def parse_rdson_vgs(ws, wb):
    """RDSon-VGS: Rdson vs Vgs, 多温度
    列: VGS, RDS(Ω), RDS(mΩ) for -55°C, then 25°C, then 150°C, T coefficient
    """
    # Row 1: VGS, RDS(Ω), RDS(mΩ), -55°C, VGS, RDS(Ω), RDS(mΩ), 25°C, ...
    # 实际: col 6=VGS(-55), 7=RDS_Ω(-55), 8=RDS_mΩ(-55), 9=VGS(25), 10=RDS_Ω(25), ...
    # Col 6,7,8: -55
    # Col 9,10,11: 25
    # Col 12,13,14: 150
    rows = []
    for r in range(2, ws.max_row + 1):
        row_vals = [c.value for c in ws[r]]
        if len(row_vals) < 15:
            continue
        # 3 个温度段
        for temp, base_col in [(-55, 6), (25, 9), (150, 12)]:
            vgs = _to_float(row_vals[base_col])
            rds_ohm = _to_float(row_vals[base_col + 1])
            rds_mohm = _to_float(row_vals[base_col + 2])
            if vgs is None or rds_ohm is None:
                continue
            rows.append({
                'vgs_v': vgs,
                'rdson_ohm': rds_ohm,
                'rdson_mohm': rds_mohm,
                'temperature_c': temp,
            })
    return rows


def parse_id_vth(ws, wb):
    """ID-VTH: 阈值特性（Id-Vg 在 Vth 附近）
    列布局类似 ID-VGS_5VDS
    """
    return parse_idvg_5vds(ws, wb)  # 复用相同结构


# ============================================================
#  主清洗流程
# ============================================================

PARSERS = {
    'ID-VGS_5VDS': ('IdVg_5Vds', parse_idvg_5vds),
    'ID-VGS_0.5VDS': ('IdVg_0p5Vds', parse_idvg_05vds),
    'ID-VDS': ('IdVd', parse_idvd),
    'VGS-Qg': ('VgsQg', parse_vgs_qg),
    'CissCossCrss-VDS': ('CvVds', parse_cv_vds),
    'IS-VSD': ('IsVsd', parse_isvsd),
    'TJ_DC': ('TjDc', parse_tj_dc),
    'RDSon-VGS': ('RdsonVgs', parse_rdson_vgs),
    'ID-VTH': ('IdVth', parse_id_vth),
}


def write_csv(rows, output_path):
    """把 list[dict] 写成 CSV"""
    if not rows:
        print(f"  [WARN] 0 rows, skip")
        return
    # 收集所有 keys
    keys = set()
    for r in rows:
        keys.update(r.keys())
    keys = sorted(keys)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(','.join(keys) + '\n')
        for r in rows:
            vals = [str(r.get(k, '')) if r.get(k) is not None else '' for k in keys]
            f.write(','.join(vals) + '\n')
    print(f"  [OK] {len(rows)} rows → {output_path.name}")


def main():
    print(f"Input: {INPUT_FILE}")
    print(f"Output: {OUTPUT_DIR}")
    print("=" * 60)
    wb = load_workbook(INPUT_FILE, data_only=True)
    report = {}
    for sheet_name, (out_name, parser) in PARSERS.items():
        if sheet_name not in wb.sheetnames:
            print(f"[SKIP] {sheet_name} not found")
            continue
        ws = wb[sheet_name]
        try:
            rows = parser(ws, wb)
            output_path = OUTPUT_DIR / f"{out_name}.csv"
            write_csv(rows, output_path)
            report[out_name] = {
                'source_sheet': sheet_name,
                'n_rows': len(rows),
                'columns': list(sorted(rows[0].keys())) if rows else [],
            }
        except Exception as e:
            print(f"  [ERROR] {sheet_name}: {e}")
            import traceback
            traceback.print_exc()
            report[out_name] = {'source_sheet': sheet_name, 'error': str(e)}

    # 写清洗报告
    report_path = INSPECT_DIR / "cleaning_report.json"
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    print(f"\nReport: {report_path}")
    print("=" * 60)
    print("Done.")


if __name__ == '__main__':
    main()

#!/usr/bin/env python3
"""
完整的 SPICE 拟合报告生成器
从数据加载、拟合到报告生成的完整流程
"""
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
from datetime import datetime

from spicebuilder.data.loader_sdh import load_sdh_excel
from spicebuilder.models.bsim3 import BSIM3Model
from spicebuilder.models.init_values import init_from_key_params
from spicebuilder.fitting.optimizer import Optimizer
from spicebuilder.strategy.sgt_6stage import build_sgt_engine
from spicebuilder.simulator.evaluator import LTspiceEvaluator


def create_header_style():
    """创建表头样式"""
    return {
        'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'font': Font(bold=True, color='FFFFFF', size=11),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def create_scatter_chart(title, x_label, y_label):
    """创建散点图（带平滑线）"""
    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = x_label
    chart.y_axis.title = y_label
    chart.style = 2
    chart.width = 16
    chart.height = 11
    return chart


def add_comparison_sheet(wb, sim_data, sheet_name, title, x_label, y_label, x_unit='', y_unit=''):
    """通用函数：创建测量值 vs 仿真值对比 Sheet"""
    ws = wb.create_sheet(sheet_name)

    # 标题
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True, color='1F4E78')
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25

    # 表头
    header_style = create_header_style()
    headers = [
        f'{x_label} ({x_unit})' if x_unit else x_label,
        f'{y_label}_measured ({y_unit})' if y_unit else f'{y_label}_measured',
        f'{y_label}_simulated ({y_unit})' if y_unit else f'{y_label}_simulated',
        'Error (%)'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        for key, value in header_style.items():
            setattr(cell, key, value)

    # 填充数据
    row = 4
    for x_val, y_meas, y_sim in zip(sim_data.ivar, sim_data.dvar, sim_data.fit):
        ws.cell(row, 1, x_val)
        ws.cell(row, 2, y_meas)
        ws.cell(row, 3, y_sim)

        # 计算相对误差
        if abs(y_meas) > 1e-12:
            error_pct = (y_sim - y_meas) / abs(y_meas) * 100
            ws.cell(row, 4, error_pct)
            ws.cell(row, 4).number_format = '0.00'

            # 根据误差大小设置颜色
            if abs(error_pct) < 5:
                ws.cell(row, 4).font = Font(color='006100')  # 绿色
            elif abs(error_pct) < 10:
                ws.cell(row, 4).font = Font(color='FF6600')  # 橙色
            else:
                ws.cell(row, 4).font = Font(color='C00000')  # 红色

        # 格式化数值
        ws.cell(row, 1).number_format = '0.000' if x_val < 100 else '0.0'

        # 根据数量级选择格式
        if abs(y_meas) < 0.01 or abs(y_meas) > 1000:
            ws.cell(row, 2).number_format = '0.000E+00'
            ws.cell(row, 3).number_format = '0.000E+00'
        else:
            ws.cell(row, 2).number_format = '0.000'
            ws.cell(row, 3).number_format = '0.000'

        row += 1

    # 创建图表
    chart = create_scatter_chart(title, f'{x_label} ({x_unit})', f'{y_label} ({y_unit})')

    # 测量值系列（蓝色圆点）
    xvalues = Reference(ws, min_col=1, min_row=4, max_row=row-1)
    y_measured = Reference(ws, min_col=2, min_row=3, max_row=row-1)
    series1 = Series(y_measured, xvalues, title_from_data=True)
    series1.marker.symbol = 'circle'
    series1.marker.size = 7
    series1.smooth = True
    series1.graphicalProperties.line.width = 25000
    chart.series.append(series1)

    # 仿真值系列（红色三角）
    y_simulated = Reference(ws, min_col=3, min_row=3, max_row=row-1)
    series2 = Series(y_simulated, xvalues, title_from_data=True)
    series2.marker.symbol = 'triangle'
    series2.marker.size = 7
    series2.smooth = True
    series2.graphicalProperties.line.width = 25000
    chart.series.append(series2)

    ws.add_chart(chart, 'F3')

    # 调整列宽
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12


def create_summary_sheet(wb, result, model, dataset):
    """创建汇总页"""
    ws = wb.active
    ws.title = 'Summary'

    # 主标题
    ws['A1'] = 'SPICE Model Fitting Report'
    ws['A1'].font = Font(size=18, bold=True, color='1F4E78')
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # 器件信息
    ws['A3'] = f'Device: {dataset.device_info.part_number}'
    ws['A3'].font = Font(size=13, bold=True)
    ws['A4'] = f'Package: {dataset.device_info.package}'
    ws['A5'] = f'BVdss Rated: {dataset.device_info.bvdss_rated_v:.0f} V'
    ws['A6'] = f'Id Rated: {dataset.device_info.id_rated_a:.0f} A'
    ws['A7'] = f'Report Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'

    # 拟合质量汇总
    ws['A9'] = 'Fitting Quality Summary'
    ws['A9'].font = Font(size=14, bold=True, color='C00000')

    header_style = create_header_style()
    headers = ['Stage', 'RMS', 'Data Points', 'Status', 'Key Parameters']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(10, col, header)
        for key, value in header_style.items():
            setattr(cell, key, value)

    # 各阶段拟合结果
    row = 11
    for sr in result.stage_results:
        ws.cell(row, 1, sr.stage_name.replace('_', ' '))
        ws.cell(row, 2, sr.rms)
        ws.cell(row, 2).number_format = '0.0000'
        ws.cell(row, 3, sr.nfev)
        ws.cell(row, 4, '✓ OK' if sr.success else '✗ FAIL')

        # 参数列表（缩短）
        params_str = ', '.join(list(sr.fitted_params.keys())[:5])
        if len(sr.fitted_params) > 5:
            params_str += '...'
        ws.cell(row, 5, params_str)

        # 根据 RMS 着色
        if sr.rms < 0.5:
            ws.cell(row, 2).font = Font(color='006100', bold=True)
        elif sr.rms < 1.0:
            ws.cell(row, 2).font = Font(color='FF6600')
        else:
            ws.cell(row, 2).font = Font(color='C00000', bold=True)
        row += 1

    # Total RMS
    ws.cell(row, 1, 'Total')
    ws.cell(row, 1).font = Font(bold=True, size=12)
    ws.cell(row, 2, result.total_rms)
    ws.cell(row, 2).number_format = '0.0000'
    ws.cell(row, 2).font = Font(bold=True, size=12, color='1F4E78')

    # 关键拟合参数
    ws.cell(row+2, 1, 'Key Fitted Parameters')
    ws.cell(row+2, 1).font = Font(size=14, bold=True, color='C00000')

    param_headers = ['Parameter', 'Value', 'Unit', 'Description']
    for col, header in enumerate(param_headers, 1):
        cell = ws.cell(row+3, col, header)
        for key, value in header_style.items():
            setattr(cell, key, value)

    key_params = [
        ('VTH0', 'V', 'Threshold Voltage'),
        ('U0', 'cm²/Vs', 'Low-field Mobility'),
        ('VSAT', 'm/s', 'Saturation Velocity'),
        ('RD', 'Ω', 'Drain Series Resistance'),
        ('RS', 'Ω', 'Source Series Resistance'),
        ('CGDO', 'F/m', 'Gate-Drain Overlap Capacitance'),
        ('CGSO', 'F/m', 'Gate-Source Overlap Capacitance'),
        ('PCLM', '', 'Channel Length Modulation'),
    ]

    param_row = row + 4
    for pname, unit, desc in key_params:
        try:
            val = model.get(pname)
            ws.cell(param_row, 1, pname)
            ws.cell(param_row, 1).font = Font(bold=True)
            ws.cell(param_row, 2, val)

            # 智能格式化
            if abs(val) < 0.001 or abs(val) > 10000:
                ws.cell(param_row, 2).number_format = '0.00E+00'
            elif abs(val) < 1:
                ws.cell(param_row, 2).number_format = '0.0000'
            else:
                ws.cell(param_row, 2).number_format = '0.00'

            ws.cell(param_row, 3, unit)
            ws.cell(param_row, 4, desc)
            param_row += 1
        except:
            pass

    # 调整列宽
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 38


def main():
    """主函数"""
    print('=' * 60)
    print('SPICE Model Fitting Report Generator')
    print('=' * 60)

    # 1. 加载数据
    print('\n[1/4] 加载测试数据...')
    dataset = load_sdh_excel('datademo/SDH10N2P1WC-AA_SPICE_Data.xlsx')
    print(f'  Device: {dataset.device_info.part_number}')
    print(f'  Id-Vg points: {len(dataset.idvg_vds5)}')
    print(f'  Id-Vd points: {len(dataset.idvd)}')

    # 2. 初始化模型并拟合
    print('\n[2/4] 运行 BSIM3 参数拟合...')
    model = BSIM3Model()
    init_from_key_params(model, dataset.key_params)

    simulator = LTspiceEvaluator(subckt_name='SDH10N2P1', rg_ohm=1.6, verbose=False)
    opt = Optimizer(method='trf')

    engine = build_sgt_engine(dataset, model, opt,
                             error_threshold=0.5, max_loops=1,
                             verbose=True, simulator=simulator)
    result = engine.run(opt)

    print(f'\n  拟合完成！Total RMS = {result.total_rms:.4f}')

    # 3. 生成报告
    print('\n[3/4] 生成 Excel 报告...')
    wb = Workbook()

    # Summary
    print('  创建 Summary...')
    create_summary_sheet(wb, result, model, dataset)

    # Id-Vg @ Vds=0.5V
    if len(engine.stages) > 0 and engine.stages[0].simdata:
        print('  创建 Id-Vg @ Vds=0.5V...')
        s1_sim = engine.stages[0].simdata[0]
        add_comparison_sheet(wb, s1_sim, 'Id-Vg_Vds0.5V',
                           'Id-Vg @ Vds=0.5V, 25°C',
                           'Vgs', 'Id', 'V', 'A')

    # Id-Vg @ Vds=5V
    if len(engine.stages) > 1 and engine.stages[1].simdata:
        print('  创建 Id-Vg @ Vds=5V...')
        s3_sim = engine.stages[1].simdata[0]
        add_comparison_sheet(wb, s3_sim, 'Id-Vg_Vds5V',
                           'Id-Vg @ Vds=5V, 25°C',
                           'Vgs', 'Id', 'V', 'A')

    # Id-Vd (多条曲线)
    if len(engine.stages) > 2:
        s4_sims = engine.stages[2].simdata
        for i, sim in enumerate(s4_sims):
            vgs = sim.metadata.get('vgs_v', 10)
            print(f'  创建 Id-Vd @ Vgs={vgs}V...')
            add_comparison_sheet(wb, sim, f'Id-Vd_Vgs{int(vgs)}V',
                               f'Id-Vd @ Vgs={vgs}V, 25°C',
                               'Vds', 'Id', 'V', 'A')

    # 4. 保存
    print('\n[4/4] 保存报告...')
    output_file = f'__FIT_REPORT__{dataset.device_info.part_number}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(output_file)

    print('\n' + '=' * 60)
    print(f'✓ 报告已生成: {output_file}')
    print(f'  Total RMS: {result.total_rms:.4f}')
    print(f'  包含 {len(wb.sheetnames)} 个 Sheet')
    print('=' * 60)


if __name__ == '__main__':
    main()

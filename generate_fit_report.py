#!/usr/bin/env python3
"""
生成完整的 SPICE 模型拟合报告
包含多个 Sheet，每个 Sheet 对应一个电学特性的测量值 vs 仿真值对比
"""
import pickle
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
from datetime import datetime


def create_header_style():
    """创建表头样式"""
    return {
        'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'font': Font(bold=True, color='FFFFFF', size=11),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def create_scatter_chart(title, x_label, y_label):
    """创建散点图（带平滑线）"""
    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = x_label
    chart.y_axis.title = y_label
    chart.style = 2
    chart.width = 16
    chart.height = 11
    return chart


def add_data_series(chart, ws, x_col, y_col, start_row, end_row, title, symbol='circle', color=None):
    """添加数据系列到图表"""
    xvalues = Reference(ws, min_col=x_col, min_row=start_row, max_row=end_row)
    yvalues = Reference(ws, min_col=y_col, min_row=start_row, max_row=end_row)

    series = Series(yvalues, xvalues, title=title)
    series.marker.symbol = symbol
    series.marker.size = 7
    series.smooth = True  # 平滑线
    series.graphicalProperties.line.width = 20000  # 2pt线宽

    chart.series.append(series)


def create_summary_sheet(wb, result, model, device_info):
    """创建汇总页"""
    ws = wb.active
    ws.title = 'Summary'

    # 标题
    ws['A1'] = 'SPICE Model Fitting Report'
    ws['A1'].font = Font(size=18, bold=True, color='1F4E78')
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 器件信息
    ws['A3'] = f'Device: {device_info.part_number}'
    ws['A3'].font = Font(size=12, bold=True)
    ws['A4'] = f'Package: {device_info.package}'
    ws['A5'] = f'BVdss: {device_info.bvdss_rated_v:.0f} V'
    ws['A6'] = f'Report Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'

    # 拟合质量汇总
    ws['A8'] = 'Fitting Quality Summary'
    ws['A8'].font = Font(size=13, bold=True, color='C00000')

    header_style = create_header_style()
    headers = ['Stage', 'RMS', 'Data Points', 'Status', 'Parameters']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(9, col, header)
        for key, value in header_style.items():
            setattr(cell, key, value)

    # 填充每个阶段的数据
    row = 10
    for sr in result.stage_results:
        ws.cell(row, 1, sr.stage_name)
        ws.cell(row, 2, sr.rms)
        ws.cell(row, 2).number_format = '0.0000'
        ws.cell(row, 3, sr.nfev)
        ws.cell(row, 4, '✓ OK' if sr.success else '✗ FAIL')
        ws.cell(row, 5, ', '.join(sr.fitted_params.keys())[:50])

        # 根据 RMS 设置颜色
        if sr.rms < 0.5:
            ws.cell(row, 2).font = Font(color='006100')  # 绿色
        elif sr.rms < 1.0:
            ws.cell(row, 2).font = Font(color='FF6600')  # 橙色
        else:
            ws.cell(row, 2).font = Font(color='C00000')  # 红色
        row += 1

    # Total RMS
    ws.cell(row, 1, 'Total')
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2, result.total_rms)
    ws.cell(row, 2).number_format = '0.0000'
    ws.cell(row, 2).font = Font(bold=True, size=12)

    # 关键参数表
    ws.cell(row+2, 1, 'Key Fitted Parameters')
    ws.cell(row+2, 1).font = Font(size=13, bold=True, color='C00000')

    param_headers = ['Parameter', 'Value', 'Unit', 'Description']
    for col, header in enumerate(param_headers, 1):
        cell = ws.cell(row+3, col, header)
        for key, value in header_style.items():
            setattr(cell, key, value)

    # 关键参数列表
    key_params = [
        ('VTH0', 'V', 'Threshold Voltage'),
        ('U0', 'cm²/Vs', 'Mobility'),
        ('VSAT', 'm/s', 'Saturation Velocity'),
        ('RD', 'Ω', 'Drain Resistance'),
        ('RS', 'Ω', 'Source Resistance'),
        ('CGDO', 'F/m', 'Gate-Drain Overlap Cap'),
        ('CGSO', 'F/m', 'Gate-Source Overlap Cap'),
    ]

    param_row = row + 4
    for pname, unit, desc in key_params:
        try:
            val = model.get(pname)
            ws.cell(param_row, 1, pname)
            ws.cell(param_row, 2, val)

            # 根据数值范围选择格式
            if abs(val) < 0.01 or abs(val) > 1000:
                ws.cell(param_row, 2).number_format = '0.00E+00'
            else:
                ws.cell(param_row, 2).number_format = '0.0000'

            ws.cell(param_row, 3, unit)
            ws.cell(param_row, 4, desc)
            param_row += 1
        except:
            pass

    # 调整列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 18


def create_idvg_sheet(wb, sim_data, sheet_name, title):
    """创建 Id-Vg 对比 Sheet"""
    ws = wb.create_sheet(sheet_name)

    # 标题
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True, color='1F4E78')
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # 表头
    header_style = create_header_style()
    headers = ['Vgs (V)', 'Id_measured (A)', 'Id_simulated (A)', 'Error (%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        for key, value in header_style.items():
            setattr(cell, key, value)

    # 数据
    row = 4
    for vgs, id_meas, id_sim in zip(sim_data.ivar, sim_data.dvar, sim_data.fit):
        ws.cell(row, 1, vgs)
        ws.cell(row, 2, id_meas)
        ws.cell(row, 3, id_sim)

        # 计算相对误差 (%)
        if id_meas > 1e-12:
            error_pct = (id_sim - id_meas) / id_meas * 100
            ws.cell(row, 4, error_pct)
            ws.cell(row, 4).number_format = '0.00'

        ws.cell(row, 1).number_format = '0.00'
        ws.cell(row, 2).number_format = '0.000E+00'
        ws.cell(row, 3).number_format = '0.000E+00'
        row += 1

    # 创建图表
    chart = create_scatter_chart(title, 'Vgs (V)', 'Id (A)')

    # 测量值系列
    add_data_series(chart, ws, 1, 2, 4, row-1, 'Measured', symbol='circle')

    # 仿真值系列
    add_data_series(chart, ws, 1, 3, 4, row-1, 'Simulated', symbol='triangle')

    ws.add_chart(chart, 'F3')

    # 调整列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12


def create_idvd_sheet(wb, sim_data, sheet_name):
    """创建 Id-Vd 对比 Sheet"""
    ws = wb.create_sheet(sheet_name)

    vgs = sim_data.metadata.get('vgs_v', 10.0)
    title = f'Id-Vd @ Vgs={vgs}V, 25°C'

    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True, color='1F4E78')
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # 表头
    header_style = create_header_style()
    headers = ['Vds (V)', 'Id_measured (A)', 'Id_simulated (A)', 'Error (%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        for key, value in header_style.items():
            setattr(cell, key, value)

    # 数据
    row = 4
    for vds, id_meas, id_sim in zip(sim_data.ivar, sim_data.dvar, sim_data.fit):
        ws.cell(row, 1, vds)
        ws.cell(row, 2, id_meas)
        ws.cell(row, 3, id_sim)

        if id_meas > 1e-3:
            error_pct = (id_sim - id_meas) / id_meas * 100
            ws.cell(row, 4, error_pct)
            ws.cell(row, 4).number_format = '0.00'

        ws.cell(row, 1).number_format = '0.000'
        ws.cell(row, 2).number_format = '0.000'
        ws.cell(row, 3).number_format = '0.000'
        row += 1

    # 创建图表
    chart = create_scatter_chart(title, 'Vds (V)', 'Id (A)')
    add_data_series(chart, ws, 1, 2, 4, row-1, 'Measured', symbol='circle')
    add_data_series(chart, ws, 1, 3, 4, row-1, 'Simulated', symbol='triangle')

    ws.add_chart(chart, 'F3')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12


def create_cv_sheet(wb, sim_data, sheet_name, cap_type):
    """创建 C-V 对比 Sheet"""
    ws = wb.create_sheet(sheet_name)

    title = f'{cap_type.upper()} vs Vds @ 25°C'

    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True, color='1F4E78')
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # 表头
    header_style = create_header_style()
    headers = ['Vds (V)', 'C_measured (pF)', 'C_simulated (pF)', 'Error (%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        for key, value in header_style.items():
            setattr(cell, key, value)

    # 数据
    row = 4
    for vds, c_meas, c_sim in zip(sim_data.ivar, sim_data.dvar, sim_data.fit):
        ws.cell(row, 1, vds)
        ws.cell(row, 2, c_meas)
        ws.cell(row, 3, c_sim)

        if c_meas > 1e-3:
            error_pct = (c_sim - c_meas) / c_meas * 100
            ws.cell(row, 4, error_pct)
            ws.cell(row, 4).number_format = '0.00'

        ws.cell(row, 1).number_format = '0.0'
        ws.cell(row, 2).number_format = '0.0'
        ws.cell(row, 3).number_format = '0.0'
        row += 1

    # 创建图表
    chart = create_scatter_chart(title, 'Vds (V)', 'C (pF)')
    add_data_series(chart, ws, 1, 2, 4, row-1, 'Measured', symbol='circle')
    add_data_series(chart, ws, 1, 3, 4, row-1, 'Simulated', symbol='triangle')

    ws.add_chart(chart, 'F3')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12


def main():
    """主函数"""
    print('正在生成 SPICE 拟合报告...')

    # 加载拟合结果
    with open('temp_fit_result.pkl', 'rb') as f:
        data = pickle.load(f)

    result = data['result']
    model = data['model']
    dataset = data['dataset']
    engine = data['engine']

    # 创建工作簿
    wb = Workbook()

    # 1. Summary Sheet
    print('  [1/8] 创建 Summary...')
    create_summary_sheet(wb, result, model, dataset.device_info)

    # 2. Id-Vg @ Vds=0.5V
    print('  [2/8] 创建 Id-Vg @ Vds=0.5V...')
    s1_sim = engine.stages[0].simdata[0]
    create_idvg_sheet(wb, s1_sim, 'Id-Vg_Vds0.5V', 'Id-Vg @ Vds=0.5V, 25°C')

    # 3. Id-Vg @ Vds=5V
    print('  [3/8] 创建 Id-Vg @ Vds=5V...')
    s3_sim = engine.stages[1].simdata[0]
    create_idvg_sheet(wb, s3_sim, 'Id-Vg_Vds5V', 'Id-Vg @ Vds=5V, 25°C')

    # 4-7. Id-Vd (多条曲线)
    s4_sims = engine.stages[2].simdata
    for i, sim in enumerate(s4_sims[:4], 4):  # 最多4条
        vgs = sim.metadata.get('vgs_v', 10)
        print(f'  [{i}/8] 创建 Id-Vd @ Vgs={vgs}V...')
        create_idvd_sheet(wb, sim, f'Id-Vd_Vgs{int(vgs)}V')

    # 8. C-V (如果有 S6)
    if len(engine.stages) > 4:
        print(f'  [8/8] 创建 C-V...')
        # S6 可能包含多个 SimData (Ciss/Coss/Crss)
        # 这里简化处理，仅取第一个
        s6_sim = engine.stages[4].simdata[0] if engine.stages[4].simdata else None
        if s6_sim and s6_sim.curve_type == 'CvVds':
            cap_type = s6_sim.metadata.get('cap_type', 'ciss')
            create_cv_sheet(wb, s6_sim, f'C-V_{cap_type.upper()}', cap_type)

    # 保存
    output_file = f'__FIT_REPORT__{dataset.device_info.part_number}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(output_file)
    print(f'\n[OK] Report generated: {output_file}')
    print(f'  Total RMS: {result.total_rms:.4f}')
    print(f'  Sheets: {len(wb.sheetnames)}')


if __name__ == '__main__':
    main()
